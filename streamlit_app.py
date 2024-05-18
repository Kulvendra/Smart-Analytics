import streamlit as st
import pandas as pd
from pathlib import Path
from io import StringIO
import json
import openai
import openpyxl

openai.api_key = "you-open-ai-key"
is_write_files =False
chart_list=[]

def render_all_echarts(all_configs):
    # Start of the HTML document
    echarts_html_template = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <script src="https://cdn.jsdelivr.net/npm/echarts/dist/echarts.min.js"></script>
        <style>
            .echart-container { width: 600px; height: 400px; margin-bottom: 20px; }
        </style>
    </head>
    <body>
    """

    # Generate a chart container and JavaScript for each chart in the list
    for chart_title, config in all_configs.items():
        echarts_html_template += f'<div id="{chart_title.replace(" ", "_")}" class="echart-container"></div>\n'
        echarts_html_template += f"""
        <script type="text/javascript">
            var myChart_{chart_title.replace(" ", "_")} = echarts.init(document.getElementById('{chart_title.replace(" ", "_")}'));
            var option = {json.dumps(config)};
            myChart_{chart_title.replace(" ", "_")}.setOption(option);
        </script>
        """

    # End of the HTML document
    echarts_html_template += "</body></html>"

    if is_write_files:
        with open("chartView.html", 'w') as file:
            file.write(echarts_html_template)
    
    return echarts_html_template


def build_echart(excel_file, sheet_name,project_description,chart_list,st_state,query):


    if 'messages' not in st_state:
        st_state['messages'] = [
            {"role": "system", "content": "You generate echart config"},
            {"role": "user", "content": f"Generate ECharts configuration JSON based on the following input chart data and include x-axis and y-axis labels. The output should be a dictionary where each key is the chart's title and the value is the ECharts configuration.Input=> {chart_list} "}
        ]
    st_state['messages'].append({"role": "system", "content": query})
    # print("Generating Echarts...")

    # prompt = [
    #     {"role": "system", "content": "You generate echart config"},
    #     {"role": "user", "content": f"Generate ECharts configuration JSON based on the following input chart data and include x-axis and y-axis labels. The output should be a dictionary where each key is the chart's title and the value is the ECharts configuration.Input=> {chart_list} "}
    #     ]
    response = openai.chat.completions.create(
    model="gpt-4o",
    response_format={ "type": "json_object" },
    messages= st_state['messages']
    )
    echarts = response.choices[0].message.content

    echarts = json.loads(echarts) 
    filename = 'echarts.json'
    if is_write_files:
        with open(filename, 'w') as f:
            json.dump(echarts, f)

    print("Echarts are generated..!!")
    return echarts


def generate_chart_list(excel_file, sheet_name,project_description):
    print("Generating chart list...")
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return []  # Return an empty list if the specified sheet does not exist
    sheet = workbook[sheet_name]  # Select the specified sheet

    headers = []
    first_few_rows = []
    max_rows_to_check = 10  # You can adjust this number based on typical file structures

    # Collect data from the first few rows to determine the header
    for row in sheet.iter_rows(min_row=1, max_row=max_rows_to_check):
        # Use list comprehension to process each cell in a row
        row_data = [(cell.value if len(str(cell.value)) <= 50 else str(cell.value)[:50]) for cell in row if cell.value is not None]
        if row_data:  # Ensure the row is not empty
            first_few_rows.append(row_data)

    # Convert rows data to a string format suitable for sending to ChatGPT
    rows_text = "\n".join([", ".join(map(str, row)) for row in first_few_rows])
    rows_text = rows_text + project_description

    content = "You are a data analyst who creates chart from input file and project description."

    prompt = '''Step 1: Analyze the dataset to identify meaningful relationships between pairs of headers, excluding any headers related to time.
                Step 2: From the analysis, list pairs of headers that can form meaningful 2D charts, focusing on statistically significant or contextually important relationships.
                Step 3: For each identified pair, suggest a suitable 2D and 3D chart type (e.g., bar, line, scatter, heatmap,treemap,sunburst,graph,tree,scatter,candlestick,radar), considering the nature of the data.Need to plot atleast two chart type in heatmap,treemap,sunburst,graph,tree,scatter,candlestick,radar with line and bar
                Step 4: Specify which headers should be used for the x and y axes for each chart, ensuring clarity and insight.
                Step 5: Provide a structured list of these chart suggestions, including the chart type, x axis, and y axis for each chart.
                
                Output Json format 
                {
                    "chartList": [
                        {
                            "x": "Matter Name",
                            "y": "Total Fees",
                            "chartType": "bar",
                            "title": "Total Fees by Matter Name",
                            "x-data": [
                                "Mary Hartman vs AiCaramba Inc.",
                                "AiCaramba v Panda Bear",
                                "James Dean",
                                "Mary Hartman vs AiCaramba Inc."
                            ],
                            "y-data": [
                                564300,
                                688160.5,
                                497186,
                                564300
                            ]
                        }
                    ]
                }
                '''
    prompt = prompt + rows_text
    response = openai.chat.completions.create(
    model="gpt-4o",
    response_format={ "type": "json_object" },
    messages=[
        {"role": "system", "content": "You generate echart config"},
        {"role": "user", "content": prompt}
    ]
    )
    config = response.choices[0].message.content

    config = json.loads(config) 
  
    filename = 'chartList.json'

    if is_write_files:
        with open(filename, 'w') as f:
            json.dump(config, f)

    print("Chart list is generated")
    return config 



def start(payload, project_description,st_state, query):
    global chart_list
    chart_list = generate_chart_list(payload['file_paths'], payload['sheet_name'], project_description)        
    echart = build_echart(payload['file_paths'], payload['sheet_name'], project_description, chart_list,st_state, query)
    html_view = render_all_echarts(echart)
    return html_view


def main():
    global chart_list
    # Streamlit App
    st.title('Echart Generator')

    st.header('Upload Your File')
    uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])

    st.header('Sheet Name')
    sheet_name = st.text_input("Enter sheet name")

    st.header('Project Description')
    project_description = st.text_area("Enter project description")

    st.header('Query')
    query = st.text_input("Enter your query")

    if 'st_state' not in st.session_state:
        st.session_state['st_state'] = {}

    generate_bt = st.button('Generate Charts')
    build_bt =st.button('Build Charts')

    if generate_bt:
        if uploaded_file and project_description :
            payload = {
                'file_paths': uploaded_file,
                'sheet_name': sheet_name  # Assuming sheet_name is optional or can be None
            }
            html_view = start(payload, project_description,st.session_state['st_state'],query)
            st.components.v1.html(html_view, height=1500, scrolling=True)
        else:
            st.error("Please upload files, enter a project description, and a query.")
            
    elif build_bt:
        json_str = json.dumps(chart_list)
        st.text("JSON as string:")
        st.text(json_str)
        # echart = build_echart(payload['file_paths'], payload['sheet_name'], project_description, chart_list,st_state, query)

    
    

if __name__ == "__main__":
    main()