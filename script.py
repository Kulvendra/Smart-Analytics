

import openai
import openpyxl
import json
openai.api_key = "you-open-ai-key"


def render_all_echarts(all_configs):
    # Start of the HTML document
    echarts_html_template = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <script src="https://cdn.jsdelivr.net/npm/echarts/dist/echarts.min.js"></script>
        <style>
            .echart-container { width: 600px; height: 400px; margin-bottom: 20px; }
        </style>
    </head>
    <body>
    """

    # Generate a chart container and JavaScript for each chart in the list
    for chart_title, config in all_configs.items():
        echarts_html_template += f'<div id="{chart_title.replace(" ", "_")}" class="echart-container"></div>\n'
        echarts_html_template += f"""
        <script type="text/javascript">
            var myChart_{chart_title.replace(" ", "_")} = echarts.init(document.getElementById('{chart_title.replace(" ", "_")}'));
            var option = {json.dumps(config)};
            myChart_{chart_title.replace(" ", "_")}.setOption(option);
        </script>
        """

    # End of the HTML document
    echarts_html_template += "</body></html>"

    with open("chartView.html", 'w') as file:
        file.write(echarts_html_template)
    
    return echarts_html_template


def build_echart(excel_file, sheet_name,project_description,chart_list,query):
    print("Generating Echarts...")

    prompt = [
        {"role": "system", "content": "You generate echart config"},
        {"role": "user", "content": f"Generate ECharts configuration JSON based on the following input chart data and include x-axis and y-axis labels. The output should be a dictionary where each key is the chart's title and the value is the ECharts configuration.Input=> {chart_list} "}
        ]
    response = openai.chat.completions.create(
    model="gpt-4o",
    response_format={ "type": "json_object" },
    messages= prompt
    )
    echarts = response.choices[0].message.content

    echarts = json.loads(echarts) 
    filename = 'echarts.json'
    with open(filename, 'w') as f:
        json.dump(echarts, f)

    print("Echarts are generated..!!")
    return echarts


def generate_chart_list(excel_file, sheet_name,project_description):
    print("Generating chart list...")
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return []  # Return an empty list if the specified sheet does not exist
    sheet = workbook[sheet_name]  # Select the specified sheet

    headers = []
    first_few_rows = []
    max_rows_to_check = 10  # You can adjust this number based on typical file structures

    # Collect data from the first few rows to determine the header
    for row in sheet.iter_rows(min_row=1, max_row=max_rows_to_check):
        # Use list comprehension to process each cell in a row
        row_data = [(cell.value if len(str(cell.value)) <= 50 else str(cell.value)[:50]) for cell in row if cell.value is not None]
        if row_data:  # Ensure the row is not empty
            first_few_rows.append(row_data)

    # Convert rows data to a string format suitable for sending to ChatGPT
    rows_text = "\n".join([", ".join(map(str, row)) for row in first_few_rows])
    rows_text = rows_text + project_description

    content = "You are a data analyst who creates chart from input file and project description."

    prompt = '''Step 1: Analyze the dataset to identify meaningful relationships between pairs of headers, excluding any headers related to time.
                Step 2: From the analysis, list pairs of headers that can form meaningful 2D charts, focusing on statistically significant or contextually important relationships.
                Step 3: For each identified pair, suggest a suitable 2D and 3D chart type (e.g., bar, line, scatter, heatmap,treemap,sunburst,graph,tree,scatter,candlestick,radar), considering the nature of the data.Need to plot atleast two chart type in heatmap,treemap,sunburst,graph,tree,scatter,candlestick,radar
                Step 4: Specify which headers should be used for the x and y axes for each chart, ensuring clarity and insight.
                Step 5: Provide a structured list of these chart suggestions, including the chart type, x axis, and y axis for each chart.
                
                Output Json format 
                {
                    "chartList": [
                        {
                            "x": "Matter Name",
                            "y": "Total Fees",
                            "chartType": "bar",
                            "title": "Total Fees by Matter Name",
                            "x-data": [
                                "Mary Hartman vs AiCaramba Inc.",
                                "AiCaramba v Panda Bear",
                                "James Dean",
                                "Mary Hartman vs AiCaramba Inc."
                            ],
                            "y-data": [
                                564300,
                                688160.5,
                                497186,
                                564300
                            ]
                        }
                    ]
                }
                '''
    prompt = prompt + rows_text
    response = openai.chat.completions.create(
    model="gpt-4o",
    response_format={ "type": "json_object" },
    messages=[
        {"role": "system", "content": "You generate echart config"},
        {"role": "user", "content": prompt}
    ]
    )
    config = response.choices[0].message.content

    config = json.loads(config) 
  
    filename = 'chartList.json'
    with open(filename, 'w') as f:
        json.dump(config, f)

    print("Chart list is generated")
    return config 



def api_smart_charts(payload,project_description,query):
    chart_list = generate_chart_list(payload['file_paths'],payload['sheet_name'],project_description)        
    ehcart = build_echart(payload['file_paths'],payload['sheet_name'],project_description,chart_list,query)
    html_view = render_all_echarts(ehcart)
    return html_view



payload = {"file_paths":"Invoice Library.xlsx","sheet_name":"InvoiceLineItems"}
project_description = '''
    Project Description :
    The project is related to invoice and matter data. It should address the following use cases or functionalities
    1. Ability to summarize key facts, issues and current status of specific matters
    2. Based on current status and complexity of a matter, estimate the potential legal costs and timeline to resolution.
    3. Provide a range of expected legal fees for specific matter types
    4. Identify potential risks with specific matters
    5. Provide document summaries
    6. Draft documents based on templates and known facts about the matter
    7. Review data produced in discovery and identify relevant information
    8. Provide proposed timelines and milestones for a matter
    9. Identify any outlier or unusual timekeeper rates charged by a firm across matter types/timekeeper levels
    10. Calculate average hourly rate charged by each firm by matter type for time period
    11. Breakdown legal spend by matter, by timekeeper, by firm for time period
    12. Identify matters/firms that are close to exceeding budget or have exceeded budget
    13. Identify firms routinely exceeding the matter budget
    14. Analyze legal spend trends
    15. Forecast legal spend based on current historical data
    16. Provide range of estimated legal spend for a specific matter type based on specific timekeeper rates and billing trends.
    17. Evaluate the efficiency and cost-effectiveness of firms in handling specific matters/matter types based on timekeeper rates, staffing and overall spend
    18. Compare performance metrics of firms
    19. Suggest strategies to optimize legal spend with a firm while maintaining quality and effectiveness.
    '''
html_content = api_smart_charts(payload,project_description)
